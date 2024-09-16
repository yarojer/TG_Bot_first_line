import requests
import openpyxl
from time import sleep
from datetime import datetime, timedelta

# Функция для генерации списка дат в заданном интервале
def generate_dates(start_date, end_date):
    current_date = start_date
    while current_date <= end_date:
        yield current_date
        current_date += timedelta(days=1)

# Загружаем книгу Excel
wb = openpyxl.load_workbook('Y:\\RK\\NAG.xlsx')
auth_sheet = wb['Authorization']
data_sheet = wb['ID Всех RK']

api_key = "eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjQwNzE1djEiLCJ0eXAiOiJKV1QifQ.eyJlbnQiOjEsImV4cCI6MTczODAwOTUyNCwiaWQiOiIzMGY4MDdkMy01OGY4LTQwYmItYmJjOS0xZGRjZTkyZDNiOGEiLCJpaWQiOjE2OTU5Mzc5LCJvaWQiOjEzNDUxNSwicyI6OTYsInNpZCI6IjBiMTBjNjdmLTMwM2MtNDc4OC05YmExLTFhZWIyYzE4MWI3OSIsInQiOmZhbHNlLCJ1aWQiOjE2OTU5Mzc5fQ.J6LWKxf6L65y3AG_HJ6qTISshkuX3NKmb20FGYfDG4Frm-uOw1SphjYJ8L4Tzit9rNADuOpBFG5xHvW8ZiJN8A"

date_from = auth_sheet['B3'].value
date_to = auth_sheet['B4'].value

# Преобразуем даты из строкового формата в объекты datetime
date_from = datetime.strptime(date_from, '%Y-%m-%d')
date_to = datetime.strptime(date_to, '%Y-%m-%d')

# Собираем IDs из листа 'ID Всех RK'
ids = [cell for row in data_sheet.iter_rows(min_row=1, max_col=1, values_only=True) for cell in row if cell is not None]

url = 'https://advert-api.wildberries.ru/adv/v2/fullstats'
headers = {
    'Authorization': api_key,
    'Content-Type': 'application/json'
}

# Создаем или очищаем лист 'Расходы по всем РК'
if 'Расходы по всем РК' not in wb.sheetnames:
    wb.create_sheet('Расходы по всем РК')
rk_sum_sheet = wb['Расходы по всем РК']
rk_sum_sheet.delete_rows(1, rk_sum_sheet.max_row)
rk_sum_sheet.append(["advertId", "nmId", "nmName", "sum"])

# Словарь для суммирования расходов по каждому ID и NM
sum_by_id_nm = {}

# Генерация списка дат в интервале
dates = list(generate_dates(date_from, date_to))
date_strings = [date.strftime('%Y-%m-%d') for date in dates]

# Отправка запросов пакетами по 50 ID и суммирование расходов
for i in range(0, len(ids), 100):
    batch_ids = ids[i:i+100]

    # Формируем пакет запросов для всего периода
    batch = [{
        "id": id,
        "dates": date_strings
    } for id in batch_ids]

    response = requests.post(url, json=batch, headers=headers)
    
    if response.status_code != 200:
        print(f"Ошибка при запросе данных: {response.status_code} {response.text}")
    else:
        data = response.json()
        print(f"Полученные данные: {data}")

        if isinstance(data, list):
            for item in data:
                advert_id = item.get('advertId')
                days = item.get('days', [])
                for day in days:
                    apps = day.get('apps', [])
                    for app in apps:
                        nms = app.get('nm', [])
                        for nm in nms:
                            nm_id = nm.get('nmId')
                            nm_name = nm.get('name')
                            expense = nm.get('sum', 0)  # Добавлено значение по умолчанию
                            key = (advert_id, nm_id, nm_name)
                            if key in sum_by_id_nm:
                                sum_by_id_nm[key] += expense
                            else:
                                sum_by_id_nm[key] = expense
        else:
            print(f"Полученный ответ не является массивом: {data}")

    sleep(65)  # Задержка в 60 секунд после каждого запроса

# Записываем данные в лист 'Расходы по всем РК'
for (advert_id, nm_id, nm_name), total_sum in sum_by_id_nm.items():
    rk_sum_sheet.append([advert_id, nm_id, nm_name, total_sum])

try:
    wb.save('Y:\\RK\\NAG.xlsx')  # Сохраняем изменения в файле Excel
    print("Файл успешно сохранен.")
except Exception as e:
    print(f"Ошибка при сохранении файла: {e}")
