import requests
import os
import time
import openpyxl
from datetime import datetime, timedelta
from dotenv import load_dotenv
import logging

# Настройка логирования
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

load_dotenv()

# Словарь для преобразования заголовков на русский язык
header_translation = {
    "date": "Дата",
    "logWarehouseCoef": "Коэффициент склада",
    "officeId": "ID Склада",
    "warehouse": "Склад",
    "warehouseCoef": "Коэффициент склада",
    "giId": "ID группы",
    "chrtId": "ID характеристики",
    "size": "Размер",
    "barcode": "Штрихкод",
    "subject": "Предмет",
    "brand": "Бренд",
    "vendorCode": "Код продавца",
    "nmId": "Артикул WB",
    "volume": "Объем",
    "calcType": "Тип расчета",
    "warehousePrice": "Стоимость хранения",
    "barcodesCount": "Количество штрихкодов",
    "palletPlaceCode": "Код паллетного места",
    "palletCount": "Количество паллет",
    "originalDate": "Оригинальная дата",
    "loyaltyDiscount": "Скидка лояльности",
    "tariffFixDate": "Дата фиксированного тарифа",
    "tariffLowerDate": "Дата снижения тарифа"
}

# Функция для получения списка всех дней в заданном диапазоне
def date_range(start_date, end_date):
    for n in range(int((end_date - start_date).days) + 1):
        yield start_date + timedelta(n)

# Функция для выполнения запросов к API WB и сохранения данных в Excel
def fetch_and_save_data(user_id, start_date, end_date, api_key):
    logger.info(f"Fetching data from {start_date} to {end_date} for user {user_id} with api_key {api_key}")

    # Создаем общий файл
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Удаляем дефолтный лист

    # Переменные для итоговых данных
    total_warehouse_price = 0
    nm_id_warehouse_prices = {}

    for single_date in date_range(start_date, end_date):
        date_str = single_date.strftime("%Y-%m-%d")
        urldata = f'https://seller-analytics-api.wildberries.ru/api/v1/paid_storage?dateFrom={date_str}&dateTo={date_str}'
        headers = {
            'accept': 'application/json',
            'Authorization': api_key
        }

        try:
            res = requests.get(urldata, headers=headers)
            res.raise_for_status()
        except requests.RequestException as e:
            logger.error(f'Ошибка при запросе данных: {e}')
            return None

        taskidres = res.json()
        taskid = taskidres["data"]["taskId"]
        urlstat = f"https://seller-analytics-api.wildberries.ru/api/v1/paid_storage/tasks/{taskid}/status"
        max_time = 400
        attempt_interval = 60
        start_time = time.time()

        # Пауза перед первым запросом статуса
        #logger.info(f"Ожидание 60 секунд перед первым запросом статуса задачи для даты {date_str}.")
        # time.sleep(60)

        while time.time() - start_time < max_time:
            try:
                res = requests.get(urlstat, headers=headers)
                res.raise_for_status()
                status = res.json().get('data', {}).get('status', '')

                if status == 'done':
                    logger.info(f'Задача выполнена на дату {date_str}.')
                    break
                else:
                    logger.info('Задача еще не выполнена. Проверим статус через 60 секунд.')
                    time.sleep(attempt_interval)  # Пауза в 60 секунд перед следующим запросом
            except requests.RequestException as e:
                logger.error(f'Ошибка при запросе статуса задачи: {e}')
                return None
        else:
            logger.warning(f'Превышено время ожидания выполнения задачи для даты {date_str}.')
            continue

        urldow = f'https://seller-analytics-api.wildberries.ru/api/v1/paid_storage/tasks/{taskid}/download'
        try:
            res = requests.get(urldow, headers=headers)
            res.raise_for_status()
            data = res.json()
        except requests.RequestException as e:
            logger.error(f'Ошибка при загрузке данных: {e}')
            return None

        if data:
            headers = list(data[0].keys())
            translated_headers = [header_translation.get(header, header) for header in headers]

            ws = wb.create_sheet(title=date_str)  # Создаем новый лист с названием даты
            ws.append(translated_headers)

            for item in data:
                row_data = [item.get(header) for header in headers]
                ws.append(row_data)

                # Подсчет суммы warehousePrice для текущей строки
                warehouse_price = item.get("warehousePrice", 0)
                total_warehouse_price += warehouse_price

                # Подсчет суммы warehousePrice по каждому nmId
                nm_id = item.get("nmId")
                if nm_id:
                    if nm_id not in nm_id_warehouse_prices:
                        nm_id_warehouse_prices[nm_id] = 0
                    nm_id_warehouse_prices[nm_id] += warehouse_price

    # Создание листа с итоговыми данными
    ws_summary = wb.create_sheet(title="Итоговые данные")
    ws_summary.append(["Показатель", "Значение"])
    ws_summary.append(["Суммарная стоимость хранения (warehousePrice)", total_warehouse_price])

    # Добавляем данные по каждому артикулу (nmId)
    ws_summary.append(["Артикул WB", "Стоимость хранения"])
    for nm_id, total_price in nm_id_warehouse_prices.items():
        ws_summary.append([nm_id, total_price])

    # Генерация уникального имени файла с меткой времени
    user_folder = f"results/{user_id}"
    os.makedirs(user_folder, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_name = f"{user_folder}/storage_{start_date.strftime('%d-%m-%Y')}_to_{end_date.strftime('%d-%m-%Y')}_{timestamp}.xlsx"
    wb.save(file_name)

    logger.info(f"Все данные успешно записаны в файл '{file_name}'")
    return file_name
