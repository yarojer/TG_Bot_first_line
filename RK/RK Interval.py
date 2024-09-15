import requests
import openpyxl
from time import sleep

# Загружаем книгу Excel
wb = openpyxl.load_workbook('X:\\RK\\ip-ryikunova-olga-aleksandrovna.xlsx')
auth_sheet = wb['Authorization']
data_sheet = wb['ID Всех RK']

api_key = "eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjQwNzE1djEiLCJ0eXAiOiJKV1QifQ.eyJlbnQiOjEsImV4cCI6MTczNjk5MjMwMSwiaWQiOiJiMTE2MWQ0Ny1iYTc3LTQ4MjgtYjg4Yy01ZWQxZjc3NzBkOGEiLCJpaWQiOjQ1ODAwNDYyLCJvaWQiOjE3MjkxLCJzIjoxMDczNzQxODg4LCJzaWQiOiI5NDM1ZjFiYy03NzFiLTU3YzEtYWUyMC1kMjlmYTNmZWRhOWUiLCJ0IjpmYWxzZSwidWlkIjo0NTgwMDQ2Mn0.7yCwqG1gknrDNGkMEcsfkAdiStj3OJv58Nl-YsR7BZEBSSO-OW8cyP94wP3_w-yck3X-JzK8stEsgFM6mDOG2g"

date_from = auth_sheet['B3'].value
date_to = auth_sheet['B4'].value

# Собираем IDs из листа 'ID Всех RK'
ids = [cell for row in data_sheet.iter_rows(min_row=1, max_col=1, values_only=True) for cell in row if cell is not None]

url = 'https://advert-api.wb.ru/adv/v2/fullstats'
headers = {
    'Authorization': api_key,
    'Content-Type': 'application/json'
}

# Создаем или очищаем лист 'Расходы по всем РК'
if 'Расходы по всем РК' not in wb.sheetnames:
    wb.create_sheet('Расходы по всем РК')
rk_sum_sheet = wb['Расходы по всем РК']
rk_sum_sheet.delete_rows(1, rk_sum_sheet.max_row)
rk_sum_sheet.append(["advertId", "sum"])

# Отправка запросов пакетами по 100 ID
for i in range(0, len(ids), 40):
    batch = [{
        "id": id,
        "interval": {
            "begin": date_from,
            "end": date_to
        }
    } for id in ids[i:i+40]]

    response = requests.post(url, json=batch, headers=headers)
    
    if response.status_code == 200:
        data = response.json()
        print(f"Получен ответ от сервера для пакета {i // 50 + 1}: {data}")
        if isinstance(data, list):
            for item in data:
                rk_sum_sheet.append([item.get('advertId'), item.get('sum')])
        else:
            print(f"Полученный ответ не является массивом: {data}")
    else:
        print(f"Ошибка при получении данных от сервера для пакета {i // 50 + 1}: {response.status_code}")

    sleep(60)  # Задержка в 60 секунд

wb.save('ip-ryikunova-olga-aleksandrovna.xlsx')  # Сохраняем изменения в файле Excel
#print("Скрипт завершил работу. Все изменения сохранены.")
