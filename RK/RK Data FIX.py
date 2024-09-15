import requests
import openpyxl
from time import sleep
from datetime import datetime, timedelta

# Функция для генерации списка дат в заданном интервале
def generate_dates(start_date, end_date):
    current_date = start_date
    while current_date <= end_date:
        yield current_date
        current_date += timedelta(days=1)

api_key = "eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjQwNzE1djEiLCJ0eXAiOiJKV1QifQ.eyJlbnQiOjEsImV4cCI6MTczNjk5MjMwMSwiaWQiOiJiMTE2MWQ0Ny1iYTc3LTQ4MjgtYjg4Yy01ZWQxZjc3NzBkOGEiLCJpaWQiOjQ1ODAwNDYyLCJvaWQiOjE3MjkxLCJzIjoxMDczNzQxODg4LCJzaWQiOiI5NDM1ZjFiYy03NzFiLTU3YzEtYWUyMC1kMjlmYTNmZWRhOWUiLCJ0IjpmYWxzZSwidWlkIjo0NTgwMDQ2Mn0.7yCwqG1gknrDNGkMEcsfkAdiStj3OJv58Nl-YsR7BZEBSSO-OW8cyP94wP3_w-yck3X-JzK8stEsgFM6mDOG2g"

date_from = auth_sheet['B3'].value
date_to = auth_sheet['B4'].value

# Преобразуем даты из строкового формата в объекты datetime
date_from = datetime.strptime(date_from, '%Y-%m-%d')
date_to = datetime.strptime(date_to, '%Y-%m-%d')

# Собираем IDs из листа 'ID Всех RK'
ids = [cell for row in data_sheet.iter_rows(min_row=1, max_col=1, values_only=True) for cell in row if cell is not None]

url = 'https://advert-api.wildberries.ru/adv/v2/fullstats'
headers = {
    'Authorization': api_key,
    'Content-Type': 'application/json'
}

# Создаем или очищаем лист 'Расходы по всем РК'
if 'Расходы по всем РК' not in wb.sheetnames:
    wb.create_sheet('Расходы по всем РК')
rk_sum_sheet = wb['Расходы по всем РК']
rk_sum_sheet.delete_rows(1, rk_sum_sheet.max_row)
rk_sum_sheet.append(["advertId", "sum"])

# Словарь для суммирования расходов по каждому ID
sum_by_id = {}

# Генерация списка дат в интервале
dates = list(generate_dates(date_from, date_to))

# Отправка запросов пакетами по 50 ID и суммирование расходов
for i in range(0, len(ids), 100):
    batch_ids = ids[i:i+100]

    # Формируем пакет запросов для каждого дня
    for date in dates:
        batch = [{
            "id": id,
            "dates": [date.strftime('%Y-%m-%d')]
        } for id in batch_ids]

        response = requests.post(url, json=batch, headers=headers)
        
        if response.status_code != 200:
            print(f"Ошибка при запросе данных: {response.status_code} {response.text}")
            continue

        data = response.json()
        print(f"Полученные данные: {data}")

        if isinstance(data, list):
            for item in data:
                advert_id = item.get('advertId')
                expense = item.get('sum', 0)  # Добавлено значение по умолчанию
                if advert_id in sum_by_id:
                    sum_by_id[advert_id] += expense
                else:
                    sum_by_id[advert_id] = expense
        else:
            print(f"Полученный ответ не является массивом: {data}")

        sleep(65)  # Задержка в 60 секунд

# Записываем данные в лист 'Расходы по всем РК'
for advert_id, total_sum in sum_by_id.items():
    rk_sum_sheet.append([advert_id, total_sum])

try:
    wb.save('X:\\RK\\ip-ryikunova-olga-aleksandrovna.xlsx')  # Сохраняем изменения в файле Excel
    print("Файл успешно сохранен.")
except Exception as e:
    print(f"Ошибка при сохранении файла: {e}")
